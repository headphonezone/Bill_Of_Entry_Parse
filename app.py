import io
import re
import tempfile
import os

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment

from boe_parser import parse_boe, build_wide_row, shipper_name

st.set_page_config(page_title="BOE to Excel (Items)", page_icon="📄", layout="wide")

st.title("📄 Bill of Entry → Excel (Item-level)")
st.caption(
    "Upload an ICEGATE-format Bill of Entry PDF. Each line item becomes "
    "one Excel row, led by BOE No and Shipper Name, with multi-licence "
    "items wrapped in a single cell."
)

uploaded = st.file_uploader("Upload BOE PDF", type=["pdf"])

if uploaded is not None:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(uploaded.read())
        tmp_path = tmp.name

    try:
        with st.spinner("Parsing BOE PDF... this can take a few seconds for 20+ page BOEs."):
            parsed = parse_boe(tmp_path)
            row = build_wide_row(parsed)
    except Exception as e:
        st.error(f"Failed to parse this PDF: {e}")
        st.stop()
    finally:
        os.unlink(tmp_path)

    n_items = len(parsed["items"])
    be_no = row.get("BOE_BE No", "")
    be_date = row.get("BOE_BE Date", "")
    total_duty = row.get("BOE_TOTAL DUTY", "")
    ass_val = row.get("BOE_TOT.ASS VAL", "")

    st.success(f"Parsed BE No **{be_no}** ({be_date}) — {n_items} line items detected.")

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("BE Number", be_no or "—")
    col2.metric("Items", n_items)
    col3.metric("Assessable Value", ass_val or "—")
    col4.metric("Total Duty", total_duty or "—")

    with st.expander("Header (Part I) fields", expanded=False):
        st.dataframe(
            pd.DataFrame(list(parsed["header"].items()), columns=["Field", "Value"]),
            use_container_width=True,
            height=300,
        )

    with st.expander("Invoice / Valuation (Part II) fields", expanded=False):
        st.dataframe(
            pd.DataFrame(list(parsed["invoice"].items()), columns=["Field", "Value"]),
            use_container_width=True,
            height=300,
        )

    with st.expander(f"Item-level details ({n_items} items)", expanded=True):
        item_rows = []
        for sno, fields in parsed["items"].items():
            r = {
                "BOE No": be_no,
                "Shipper Name": shipper_name(parsed),
                "Item #": sno,
            }
            r.update(fields)
            item_rows.append(r)
        items_df = pd.DataFrame(item_rows)
        st.dataframe(items_df, use_container_width=True, height=400)

    licence_rows = []
    for lic in parsed["licences"]:
        r = {"BOE No": be_no, "Shipper Name": shipper_name(parsed), "Item #": lic.get("ItemSN", "")}
        r.update({k: v for k, v in lic.items() if k != "ItemSN"})
        licence_rows.append(r)
    licence_df = pd.DataFrame(licence_rows)

    with st.expander(f"Licence details ({len(licence_df)} licences)", expanded=False):
        st.dataframe(licence_df, use_container_width=True, height=300)

    # Excel export - Items sheet + a Licence Details sheet (one row per
    # licence, as in the PDF, instead of the ";"-joined single cell used
    # on the Items sheet).
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        items_df.to_excel(writer, index=False, sheet_name="Items")
        licence_df.to_excel(writer, index=False, sheet_name="Licence Details")

        # Cells that contain embedded newlines get wrap text so they render
        # as stacked lines instead of one run-on line.
        for sheet_name, df in (("Items", items_df), ("Licence Details", licence_df)):
            ws = writer.sheets[sheet_name]
            max_lines = 1
            for col_idx, col_name in enumerate(df.columns, start=1):
                if df[col_name].astype(str).str.contains("\n").any():
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                        max_lines = max(max_lines, str(cell.value or "").count("\n") + 1)
            if max_lines > 1:
                for row_idx in range(2, ws.max_row + 1):
                    ws.row_dimensions[row_idx].height = 15 * max_lines
    buf.seek(0)

    safe_be = re.sub(r"[^A-Za-z0-9_-]", "_", str(be_no) or "BOE")
    st.download_button(
        label="⬇️ Download Excel (item-level)",
        data=buf,
        file_name=f"BOE_{safe_be}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

else:
    st.info("Upload a BOE PDF to begin.")
